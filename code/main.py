import boto3
import os
import io
import sys
import openpyxl
import pandas as pd
from tempfile import NamedTemporaryFile

class AutomateExcelLoad:
	def __init__(self):
		self.s3_bucket = os.environ.get('MODEL_BUCKET', 'artifacts')
		self.s3_model_folder = os.environ.get('MODEL_FOLDER', 'models')
		self.s3_model_filename = os.environ.get('MODEL_FILENAME', 'filename.xlsx')
		self.s3_temp_folder = 'temp'
		self.s3_model_path = f's3://{self.s3_bucket}/{self.s3_model_folder}/{self.s3_model_filename}'
		self.s3_model_key = f'{self.s3_model_folder}/{self.s3_model_filename}'
		self.s3_temp_path = f's3://{self.s3_bucket}/temp/{self.s3_model_filename}'
		self.s3_temp_key = f'temp/{self.s3_model_filename}'
		self.query_filename = os.environ.get('QUERY_FILENAME', 'query.csv')

		self.region_name = os.environ.get('REGION_NAME', 'us-east-1')
		self.s3_resource = boto3.resource('s3', region_name=self.region_name)
		self.s3_client = boto3.client('s3', region_name=self.region_name)

		self.start_row = int(os.environ.get('START_ROW', '2'))
		self.start_column = int(os.environ.get('START_COLUMN', '1'))
		self.last_column = int(os.environ.get('LAST_COLUMN', '100'))
		self.data_name_sheet = os.environ.get('DATA_NAME_SHEET', 'DATA')

	def get_df_from_csv(self):
		response = self.s3_resource \
					.Bucket(self.s3_bucket) \
					.Object(key='query-results/' + self.query_filename) \
					.get()
		result = pd.read_csv(io.BytesIO(response['Body'].read()), encoding='utf8')
		return result

	def save_file_on_bucket(self, wb):
		self.s3_client.put_object(Bucket=self.s3_bucket, Key=self.s3_temp_key)

		with NamedTemporaryFile() as tmp:
			temp_file = f'tmp/tmp.xlsx'
			wb.save(temp_file)
			self.s3_resource.Bucket(self.s3_bucket).upload_file(Filename=temp_file, Key=self.s3_temp_key)

		print(f'Created temp file {self.s3_model_filename} in {self.s3_temp_path}\n')

	def load_data(self, wb):
		df = self.get_df_from_csv()

		sheet_ranges = wb[self.data_name_sheet]
		last_index = len(df.index) + 2
		last_column = len(df.columns)

		# Limpa todas as células do modelo
		for cell in range(self.start_row, 100):
			column_count = 0
			while column_count <= last_column:
				sheet_ranges.cell(row=cell, column=self.start_column+column_count).value = ''
				column_count += 1

		row_count = 0

		# Carrega dados
		for cell in range(self.start_row, last_index):
			column_count = 0
			while column_count < last_column:
				sheet_ranges.cell(row=cell, column=self.start_column+column_count).value = df.iloc[ row_count , column_count ]
				column_count += 1
			row_count += 1

		pivot_sheet = wb['TD']
		pivot = pivot_sheet._pivots[0]
		pivot.cache.refreshOnLoad = True

		self.save_file_on_bucket(wb)


if __name__ == "__main__":
	automate_excel_load = AutomateExcelLoad()

	model_object = automate_excel_load.s3_resource.Bucket(automate_excel_load.s3_bucket).Object(key=automate_excel_load.s3_model_key).get()
	model_path = io.BytesIO(model_object['Body'].read())
	workbook = openpyxl.load_workbook(model_path)

	automate_excel_load.load_data(workbook)
	workbook.close()